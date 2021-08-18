import os
import openpyxl # For Excel manipulation
import boto3
import traceback
from shutil import copyfile
from PIL import Image, ImageOps
from preview_generator.manager import PreviewManager

# {
#   "dest_bucket": "lambda-generate-preview-image-destination",
#   "dest_path": "pig_#{width}_#{height}.png",
#   "error_sns": "arn:aws:sns:us-west-2:020275857710:file-previewer-failed-dev-neil",
#   "format": "png",
#   "ok_sns": "arn:aws:sns:us-west-2:020275857710:file-previewer-succeeded-dev-neil",
#   "src_bucket": "lambda-generate-preview-image-source",
#   "src_path": "source_folder/pig.jpg",
#   "request_id": "23453rjdkwfhdskjfh29er82349rewfds",
#   "dimensions": [
#     {
#       "width": "1000",
#       "height": "500"
#     },
#     {
#       "width": "500",
#       "height": "250"
#     },
#     {
#       "width": "250",
#       "height": "120"
#     }
#   ]
# }

def lambda_handler(event, context):

  ### Inputs
  print("#### Received event is: " + str(event))
  src_bucket = event["src_bucket"]
  src_path = event["src_path"]
  dest_bucket = event["dest_bucket"]
  dest_path = event["dest_path"]
  dimensions = event["dimensions"]
  output_format = event["format"]
  ok_sns = event["ok_sns"]
  error_sns = event["error_sns"]
  request_id = event["request_id"]

  try:
    ### Remove the leading slash from src_path and dest_path
#     if src_path[0] == "/":
#       scr_path = src_path[1:]
#     if dest_path[0] == "/":
#       dest_path = dest_path[1:]

    ### Get the maximum size from the dimensions received
    maximum_width = 0
    maximum_height = 0
    for dimension in dimensions:
      if int(dimension["width"]) > maximum_width:
        maximum_width = int(dimension["width"])
        maximum_height = int(dimension["height"])

    ### Copy the file from S3:
    print("Copy file from source S3")
    if src_bucket != "":
      client = boto3.client('s3')
      source_file = "/tmp/" + src_path.split("/")[-1]
      client.download_file(src_bucket, src_path, source_file)
    else:
      # If the src_bucket == 0 we will not copy from S3, and use local path.
      # This is for local testing
      source_file = src_path
    print("Successfully copied file from source S3")

    ### If file is Excel and it is big, make it smaller.
    if ".xls" in src_path and os.path.getsize(source_file) > 5000000: #5MB
      print("It seems to be a big Excel file. Try to keep only first 100 lines and columns")
      # opening the source excel file
      wb1 = openpyxl.load_workbook(source_file)
      ws1 = wb1.worksheets[0]

      # opening the destination excel file
      truncated_file = source_file + ".truncated." + source_file.split(".")[-1]

      # Create new empty file
      wb = openpyxl.Workbook()
      ws =  wb.active
      ws.title = "Changed Sheet"
      wb.save(filename = truncated_file)

      wb2 = openpyxl.load_workbook(truncated_file)
      ws2 = wb2.active

      # We will write only the first 100 lines and columns
      mr = 100
      mc = 100

      # copying the cell values from source
      # excel file to destination excel file
      for i in range (1, mr + 1):
          for j in range (1, mc + 1):
              # reading cell value from source excel file
              c = ws1.cell(row = i, column = j)

              # writing the read value to destination excel file
              ws2.cell(row = i, column = j).value = c.value

      # saving the destination excel file
      print("Saving truncated file to: " + truncated_file)
      wb2.save(str(truncated_file))
      print("Saved")
      source_file = truncated_file
      print("Successfuly truncated the big Excel file")


    ### Generate preview
    print("Generating preview")
    manager = PreviewManager("/tmp/cache", create_folder= True)
    preview_image = manager.get_jpeg_preview(source_file, width=maximum_width, height=maximum_height)
    print("Successfully generated preview")

    ### Make sure the image has the expected size and format
    print("Resize and change the format")
    # Open the preview image
    img = Image.open(preview_image)

    # Calculate the differences between actual sizes and wanted sizes
    delta_w = maximum_width - img.width
    delta_h = maximum_height - img.height

    # Then actual sizes is odd (not even), the we need to add one more pixel
    precision_pixel_w = maximum_width - img.width - int(delta_w/2) - int(delta_w-(delta_w/2))
    precision_pixel_h = maximum_height - img.height - int(delta_h/2) - int(delta_h-(delta_h/2))

    # Resize the image
    ltrb_border=(int(delta_w/2), int(delta_h/2), int(delta_w-(delta_w/2)) + precision_pixel_w, int(delta_h-(delta_h/2)) + precision_pixel_h)
    img_with_border = ImageOps.expand(img, border=ltrb_border, fill='white')
    output_file = "/tmp/output_preview." + output_format
    img_with_border.convert('RGB').save(output_file)
    print("Successfully resized and formatted for the maximum (" + str(maximum_width) + ":" + str(maximum_height) + ") wanted size")

    ### Copy to S3
    print("Copy the maximum (" + str(maximum_width) + ":" + str(maximum_height) + ") preview to destination bucket")
    if dest_bucket != "":
      client = boto3.client('s3')
      client.upload_file(output_file, dest_bucket, dest_path.replace("#{width}", str(maximum_width)).replace("#{height}", str(maximum_height)))
    else:
      # If the dest_bucket == 0 we will not copy from S3, but will copy on local path
      copyfile(output_file, dest_path.replace("#{width}", str(maximum_width)).replace("#{height}", str(maximum_height)))
    print("Successfully copied the maximum (" + str(maximum_width) + ":" + str(maximum_height) + ") preview to destination bucket")

    ### Transform, and copy all other wanted dimensions to S3 bucket
    # TODO
    for dimension in dimensions:
      # Get dimensions
      current_width = int(dimension["width"])
      current_height = int(dimension["height"])

      # Skip if the dimension is the maximum one
      if current_width == maximum_width and current_height == maximum_height:
        continue

      # Open image
      image = Image.open(output_file)
      new_image = image.resize((current_width, current_height))
      output_file_resized = "/tmp/output_preview_resized." + output_format
      new_image.save(output_file_resized)

      ### Copy to S3
      print("Copy the dimension (" + str(current_width) + ":" + str(current_height) + ") to destination bucket")
      if dest_bucket != "":
        client = boto3.client('s3')
        client.upload_file(output_file_resized, dest_bucket, dest_path.replace("#{width}", str(current_width)).replace("#{height}", str(current_height)))
      else:
        # If the dest_bucket == 0 we will not copy from S3, but will copy on local path
        copyfile(output_file_resized, dest_path.replace("#{width}", str(current_width)).replace("#{height}", str(current_height)))
      print("Successfully copied the dimension (" + str(current_width) + ":" + str(current_height) + ") preview to destination bucket")

    ### Send ok message to ok SNS
    print("Sending OK to SNS")
    if src_bucket != "":
      # If the src_bucket is not empty, then the function is running locally for dev purposes, so no need to send SNS
      # Compose the OK JSON
      okJSON = {
        "success": "true",
        "request_id": request_id,
        "src_bucket": src_bucket,
        "src_path": src_path
      }
      client = boto3.client('sns')
      response = client.publish(
          TopicArn=ok_sns,
          Message=str(okJSON)
      )
    print("Successfully send OK to SNS")
    return "OK"

  except Exception as e:
    print("Error: {}".format(e))
    tracebackError = traceback.format_exc()
    print(tracebackError)
    ### Send error message to error SNS
    print("Sending ERROR to SNS")
    # Compose the ERROR JSON
    errorJSON = {
      "success": "false",
      "request_id": request_id,
      "src_bucket": src_bucket,
      "src_path": src_path
    }
    if src_bucket != "":
      # If the src_bucket is not empty, then the function is running locally for dev purposes, so no need to send SNS
      client = boto3.client('sns')
      response = client.publish(
          TopicArn=error_sns,
          Message=str(errorJSON)
      )
    print("Successfully send ERROR to SNS")
    return "ERROR"
